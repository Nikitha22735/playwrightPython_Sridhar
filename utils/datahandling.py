import csv

from openpyxl import load_workbook


def csvData(path):
    with open(path, newline="") as data:
        return list(csv.DictReader(data))


def excelData(path, sheet_name):
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook[sheet_name]
        return list(sheet.iter_rows(min_row=2, values_only=True))
    finally:
        workbook.close()


def clearExcelData(path, sheet_name):
    workbook = load_workbook(path)
    try:
        sheet = workbook[sheet_name]
        sheet.delete_rows(2, sheet.max_row)
        workbook.save(path)
    finally:
        workbook.close()
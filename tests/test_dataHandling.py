import csv
import json
import os

from dotenv import load_dotenv
from openpyxl import load_workbook
import pytest


def test_jsonHandling():
    with open("testdata/creds.json") as data:
        formattedData = json.load(data)
        print(formattedData["password"])


def test_csvHandling():
     with open("testdata\\credentails.csv") as data:
            formattedData = csv.DictReader(data)
            values=[]
            for i in formattedData:
                 values.append(i)

            print(values[1]["username"])

# pip install openpyxl
# @pytest.mark.dh
def test_excel():
     workbook = load_workbook("testdata\\sample_creds.xlsx")
     sheet = workbook["Sheet2"]
     values=[]
     for i in sheet.iter_rows(min_row=2, values_only=True):
          values.append(i)
     print(values)


def test_excel_write():
     workbook = load_workbook("testdata\\sample_creds.xlsx")
     sheet = workbook["sheet1"]
     # sheet.append(["test","test"])
     # sheet["A5"]="test"
     sheet.delete_rows(2,sheet.max_row)
     workbook.save("testdata\\sample_creds.xlsx")

# @pytest.mark.dh
def test_cli():
     print(os.getenv('usname_s'))
     print(os.getenv('pw_s'))


# pip install python-dotenv
@pytest.mark.dh
def test_cli():
     load_dotenv(os.getenv("file"))
     print(os.getenv('urls_env'))
     print(os.getenv('us_env'))
     print(os.getenv('pw_env'))
     


